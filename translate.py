from googletrans import Translator
from openpyxl import load_workbook


my_excel = 'IA10-TIAProjectTexts.xlsx'
translator = Translator()

'''#first read and save to get rid of strange formating in this particular excel
#without this changes arent visable
wb = load_workbook(filename = my_excel)
wb.save(my_excel)'''

wb = load_workbook(my_excel) 
sheet = wb.active

#iterete through rows
for i in range(2, sheet.max_row+1):
    F = sheet[f'F{str(i)}']
    G = sheet[f'G{str(i)}']
    H = sheet[f'H{str(i)}']

    en = F.value
    pl = G.value
    ro = H.value

    # if eng empty, translate form pl to eng
    if en is None and pl is not None:
        pl_to_en = translator.translate(pl, dest='en')
        sheet[f'F{str(i)}'] = pl_to_en.text
        en = sheet[f'F{str(i)}'].value
    
    #translate eng to ro
    en_to_ro = translator.translate(en, dest='ro')
    sheet[f'H{str(i)}'] = en_to_ro.text

wb.save(my_excel)